import requests
from openpyxl import load_workbook

file = input("Enter Path of the Excel File (including file extention): ")#prompting user to input file path

workbook = load_workbook(file)

sheet = workbook.active

#looping through every url in file
for x in range(sheet.max_row - 1):
	toBeRedirectedURL = (sheet.cell(row = x + 2, column = 1).value)

	#tracking redirect using requests module
	r = requests.get(toBeRedirectedURL)
	redirectedURL = r.url
	strippedURL = redirectedURL.split("?", 1)[0]

	#writing redirect back into excel file
	cell = str("B" + str(x + 2))
	sheet[cell] = strippedURL
	print(strippedURL)

workbook.save(file)
print("FINISHED")
